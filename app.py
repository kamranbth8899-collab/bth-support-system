from flask import Flask, render_template, request, redirect, session, send_from_directory, send_file
import sqlite3
import os
import pandas as pd
from datetime import datetime
import random
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "bth_secret_key"

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def add_column_if_missing(cursor, table, column, definition):
    cursor.execute(f"PRAGMA table_info({table})")
    columns = [col[1] for col in cursor.fetchall()]

    if column not in columns:
        cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def create_database():
    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS warranty_registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_number TEXT,
            status TEXT DEFAULT 'Pending',
            full_name TEXT,
            phone TEXT,
            email TEXT,
            country TEXT,
            product_category TEXT,
            product_model TEXT,
            serial_number TEXT,
            purchase_date TEXT,
            distributor TEXT,
            created_date TEXT,
            updated_date TEXT,
            completion_date TEXT,
            internal_notes TEXT,
            service_timeline TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS after_sales_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_number TEXT,
            status TEXT DEFAULT 'Pending',
            customer_name TEXT,
            email TEXT,
            phone TEXT,
            product_model TEXT,
            serial_number TEXT,
            purchase_channel TEXT,
            issue_type TEXT,
            issue_description TEXT,
            notes TEXT,
            created_date TEXT,
            updated_date TEXT,
            completion_date TEXT,
            internal_notes TEXT,
            service_timeline TEXT
        )
    """)

    for table in ["warranty_registrations", "after_sales_requests"]:
        add_column_if_missing(cursor, table, "status", "TEXT DEFAULT 'Pending'")
        add_column_if_missing(cursor, table, "created_date", "TEXT")
        add_column_if_missing(cursor, table, "updated_date", "TEXT")
        add_column_if_missing(cursor, table, "completion_date", "TEXT")
        add_column_if_missing(cursor, table, "internal_notes", "TEXT")
        add_column_if_missing(cursor, table, "service_timeline", "TEXT")

    add_column_if_missing(cursor, "after_sales_requests", "purchase_channel", "TEXT")

    connection.commit()
    connection.close()


create_database()


def save_uploaded_file_to_case_folder(case_number, uploaded_file, label):
    if uploaded_file and uploaded_file.filename != "":
        case_folder = os.path.join(app.config["UPLOAD_FOLDER"], case_number)
        os.makedirs(case_folder, exist_ok=True)

        extension = os.path.splitext(uploaded_file.filename)[1]
        filename = f"{case_number}_{label}{extension}"
        file_path = os.path.join(case_folder, filename)
        uploaded_file.save(file_path)


def save_multiple_uploaded_files_to_case_folder(case_number, uploaded_files, label):
    case_folder = os.path.join(app.config["UPLOAD_FOLDER"], case_number)
    os.makedirs(case_folder, exist_ok=True)

    count = 1
    for uploaded_file in uploaded_files:
        if uploaded_file and uploaded_file.filename != "":
            extension = os.path.splitext(uploaded_file.filename)[1]
            filename = f"{case_number}_{label}_{count}{extension}"
            file_path = os.path.join(case_folder, filename)
            uploaded_file.save(file_path)
            count += 1


def get_case_attachments(case_number):
    attachments = []

    case_folder = os.path.join(app.config["UPLOAD_FOLDER"], case_number)

    if os.path.exists(case_folder):
        for file in os.listdir(case_folder):
            attachments.append(f"{case_number}/{file}")

    # Keep old root-level attachments visible too
    if os.path.exists(app.config["UPLOAD_FOLDER"]):
        for file in os.listdir(app.config["UPLOAD_FOLDER"]):
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], file)
            if os.path.isfile(file_path) and file.startswith(case_number):
                attachments.append(file)

    return attachments


def format_excel_file(file_path):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path)
    sheet = workbook.active

    blue_fill = PatternFill(start_color="151F7A", end_color="151F7A", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        sheet.column_dimensions[column_letter].width = min(max_length + 5, 55)

    sheet.freeze_panes = "A2"
    workbook.save(file_path)


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/products")
def products():
    return render_template("products.html")


@app.route("/support")
def support():
    return render_template("support.html")

@app.route("/support/<product>/<topic>")
def support_topic(product, topic):

    product_map = {
        "pottery-wheel": "Pottery Wheel",
        "kiln": "Kiln",
        "slab-roller": "Slab Roller"
    }

    topic_map = {
        "installation": "Installation",
        "troubleshooting": "Troubleshooting",
        "downloads": "Downloads",
        "faq": "FAQ"
    }

    product_name = product_map.get(product, product)
    topic_name = topic_map.get(topic, topic)

    return render_template(
        "support-topic.html",
        product=product,
        topic=topic,
        product_name=product_name,
        topic_name=topic_name
    )
@app.route("/contact")
def contact():
    return render_template("contact.html")


@app.route("/warranty", methods=["GET", "POST"])
def warranty():
    if request.method == "POST":
        now = get_now()

        full_name = request.form.get("full_name")
        phone = request.form.get("phone")
        email = request.form.get("email")
        country = request.form.get("country")
        product_category = request.form.get("product_category")
        product_model = request.form.get("product_model")
        serial_number = request.form.get("serial_number")
        purchase_date = request.form.get("purchase_date")
        distributor = request.form.get("distributor")

        connection = sqlite3.connect("customers.db")
        cursor = connection.cursor()

        cursor.execute("""
            INSERT INTO warranty_registrations (
                status, full_name, phone, email, country,
                product_category, product_model, serial_number,
                purchase_date, distributor, created_date,
                updated_date, completion_date, internal_notes, service_timeline
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Pending", full_name, phone, email, country,
            product_category, product_model, serial_number,
            purchase_date, distributor, now, now, "", "", ""
        ))

        today = datetime.now().strftime("%Y%m%d")
        random_number = random.randint(100, 999)
        case_number = f"W-{today}-{random_number}"

        registration_id = cursor.lastrowid

        cursor.execute("""
            UPDATE warranty_registrations
            SET case_number = ?
            WHERE id = ?
        """, (case_number, registration_id))

        connection.commit()
        connection.close()

        save_uploaded_file_to_case_folder(case_number, request.files.get("nameplate_photo"), "nameplate")
        save_uploaded_file_to_case_folder(case_number, request.files.get("invoice"), "invoice")

        return render_template("success.html", case_number=case_number)

    return render_template("warranty.html")


@app.route("/after-sales", methods=["GET", "POST"])
def after_sales():
    if request.method == "POST":
        now = get_now()

        customer_name = request.form.get("customer_name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        product_model = request.form.get("product_model")
        serial_number = request.form.get("serial_number")
        purchase_channel = request.form.get("purchase_channel")
        issue_type = request.form.get("issue_type")
        issue_description = request.form.get("issue_description")
        notes = request.form.get("notes")

        connection = sqlite3.connect("customers.db")
        cursor = connection.cursor()

        cursor.execute("""
            INSERT INTO after_sales_requests (
                status, customer_name, email, phone,
                product_model, serial_number, purchase_channel,
                issue_type, issue_description, notes, created_date,
                updated_date, completion_date, internal_notes, service_timeline
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Pending", customer_name, email, phone,
            product_model, serial_number, purchase_channel,
            issue_type, issue_description, notes, now, now, "", "", ""
        ))

        today = datetime.now().strftime("%Y%m%d")
        random_number = random.randint(100, 999)
        case_number = f"AS-{today}-{random_number}"

        request_id = cursor.lastrowid

        cursor.execute("""
            UPDATE after_sales_requests
            SET case_number = ?
            WHERE id = ?
        """, (case_number, request_id))

        connection.commit()
        connection.close()

        save_multiple_uploaded_files_to_case_folder(case_number, request.files.getlist("photos"), "photo")
        save_multiple_uploaded_files_to_case_folder(case_number, request.files.getlist("videos"), "video")

        return render_template("success.html", case_number=case_number)

    return render_template("after-sales.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if username == "admin" and password == "bth123":
            session["logged_in"] = True
            return redirect("/admin")
        else:
            error = "Invalid username or password"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.route("/update-status", methods=["POST"])
def update_status():
    if "logged_in" not in session:
        return redirect("/login")

    table = request.form.get("table")
    case_id = request.form.get("case_id")
    status = request.form.get("status")
    now = get_now()

    completion_date = ""
    if status in ["Solved", "Closed"]:
        completion_date = now

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    if table == "warranty":
        cursor.execute("""
            UPDATE warranty_registrations
            SET status = ?,
                updated_date = ?,
                completion_date = CASE
                    WHEN ? != '' THEN ?
                    ELSE completion_date
                END
            WHERE id = ?
        """, (status, now, completion_date, completion_date, case_id))

    elif table == "after_sales":
        cursor.execute("""
            UPDATE after_sales_requests
            SET status = ?,
                updated_date = ?,
                completion_date = CASE
                    WHEN ? != '' THEN ?
                    ELSE completion_date
                END
            WHERE id = ?
        """, (status, now, completion_date, completion_date, case_id))

    connection.commit()
    connection.close()

    return redirect("/admin")


@app.route("/edit-case/<case_type>/<case_number>", methods=["GET", "POST"])
def edit_case(case_type, case_number):
    if "logged_in" not in session:
        return redirect("/login")

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    if request.method == "POST":
        now = get_now()

        if case_type == "warranty":
            full_name = request.form.get("full_name")
            phone = request.form.get("phone")
            email = request.form.get("email")
            country = request.form.get("country")
            product_category = request.form.get("product_category")
            product_model = request.form.get("product_model")
            serial_number = request.form.get("serial_number")
            purchase_date = request.form.get("purchase_date")
            distributor = request.form.get("distributor")

            cursor.execute("""
                UPDATE warranty_registrations
                SET full_name = ?,
                    phone = ?,
                    email = ?,
                    country = ?,
                    product_category = ?,
                    product_model = ?,
                    serial_number = ?,
                    purchase_date = ?,
                    distributor = ?,
                    updated_date = ?
                WHERE case_number = ?
            """, (
                full_name, phone, email, country, product_category,
                product_model, serial_number, purchase_date, distributor,
                now, case_number
            ))

        elif case_type == "after-sales":
            customer_name = request.form.get("customer_name")
            email = request.form.get("email")
            phone = request.form.get("phone")
            product_model = request.form.get("product_model")
            serial_number = request.form.get("serial_number")
            purchase_channel = request.form.get("purchase_channel")
            issue_type = request.form.get("issue_type")
            issue_description = request.form.get("issue_description")
            notes = request.form.get("notes")

            cursor.execute("""
                UPDATE after_sales_requests
                SET customer_name = ?,
                    email = ?,
                    phone = ?,
                    product_model = ?,
                    serial_number = ?,
                    purchase_channel = ?,
                    issue_type = ?,
                    issue_description = ?,
                    notes = ?,
                    updated_date = ?
                WHERE case_number = ?
            """, (
                customer_name, email, phone, product_model, serial_number,
                purchase_channel, issue_type, issue_description, notes,
                now, case_number
            ))

        connection.commit()
        connection.close()

        return redirect("/admin?updated=1")

    if case_type == "warranty":
        cursor.execute("""
            SELECT
                id, case_number, status, full_name, phone, email,
                country, product_category, product_model, serial_number,
                purchase_date, distributor, created_date, updated_date,
                completion_date, internal_notes, service_timeline
            FROM warranty_registrations
            WHERE case_number = ?
        """, (case_number,))
        case_data = cursor.fetchone()

    else:
        cursor.execute("""
            SELECT
                id, case_number, status, customer_name, email, phone,
                product_model, serial_number, purchase_channel, issue_type,
                issue_description, notes, created_date, updated_date,
                completion_date, internal_notes, service_timeline
            FROM after_sales_requests
            WHERE case_number = ?
        """, (case_number,))
        case_data = cursor.fetchone()

    connection.close()

    return render_template(
        "edit-case.html",
        case_type=case_type,
        case_number=case_number,
        case_data=case_data
    )


@app.route("/delete-case", methods=["POST"])
def delete_case():
    if "logged_in" not in session:
        return redirect("/login")

    case_type = request.form.get("case_type")
    case_number = request.form.get("case_number")

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    if case_type == "warranty":
        cursor.execute("""
            DELETE FROM warranty_registrations
            WHERE case_number = ?
        """, (case_number,))

    elif case_type == "after-sales":
        cursor.execute("""
            DELETE FROM after_sales_requests
            WHERE case_number = ?
        """, (case_number,))

    connection.commit()
    connection.close()

    return redirect("/admin?deleted=1")


@app.route("/update-notes", methods=["POST"])
def update_notes():
    if "logged_in" not in session:
        return redirect("/login")

    case_type = request.form.get("case_type")
    case_number = request.form.get("case_number")
    internal_notes = request.form.get("internal_notes")
    now = get_now()

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    if case_type == "warranty":
        cursor.execute("""
            UPDATE warranty_registrations
            SET internal_notes = ?,
                updated_date = ?
            WHERE case_number = ?
        """, (internal_notes, now, case_number))

    elif case_type == "after-sales":
        cursor.execute("""
            UPDATE after_sales_requests
            SET internal_notes = ?,
                updated_date = ?
            WHERE case_number = ?
        """, (internal_notes, now, case_number))

    connection.commit()
    connection.close()

    return redirect(f"/case/{case_type}/{case_number}")


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    if "logged_in" not in session:
        return redirect("/login")

    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/case/<case_type>/<case_number>")
def case_detail(case_type, case_number):
    if "logged_in" not in session:
        return redirect("/login")

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    case_data = None

    if case_type == "warranty":
        cursor.execute("""
            SELECT
                id, case_number, status, full_name, phone, email,
                country, product_category, product_model, serial_number,
                purchase_date, distributor, created_date, updated_date,
                completion_date, internal_notes, service_timeline
            FROM warranty_registrations
            WHERE case_number = ?
        """, (case_number,))
        case_data = cursor.fetchone()

    elif case_type == "after-sales":
        cursor.execute("""
            SELECT
                id, case_number, status, customer_name, email, phone,
                product_model, serial_number, purchase_channel, issue_type,
                issue_description, notes, created_date, updated_date,
                completion_date, internal_notes, service_timeline
            FROM after_sales_requests
            WHERE case_number = ?
        """, (case_number,))
        case_data = cursor.fetchone()

    connection.close()

    attachments = get_case_attachments(case_number)

    return render_template(
        "case-detail.html",
        case_type=case_type,
        case_number=case_number,
        case_data=case_data,
        attachments=attachments
    )


@app.route("/export/warranty")
def export_warranty():
    if "logged_in" not in session:
        return redirect("/login")

    connection = sqlite3.connect("customers.db")

    df = pd.read_sql_query("""
        SELECT
            id, case_number, status, full_name, phone, email,
            country, product_category, product_model, serial_number,
            purchase_date, distributor, created_date, updated_date,
            completion_date, internal_notes
        FROM warranty_registrations
    """, connection)

    connection.close()

    df.columns = [
        "ID", "Case Number", "Status", "Full Name", "Phone", "Email",
        "Country", "Product Category", "Product Model", "Serial Number",
        "Purchase Date", "Distributor", "Created Date", "Updated Date",
        "Completion Date", "Internal Notes"
    ]

    export_path = f"warranty_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    df.to_excel(export_path, index=False)
    format_excel_file(export_path)

    return send_file(export_path, as_attachment=True)


@app.route("/export/after-sales")
def export_after_sales():
    if "logged_in" not in session:
        return redirect("/login")

    connection = sqlite3.connect("customers.db")

    df = pd.read_sql_query("""
        SELECT
            id, case_number, status, customer_name, email, phone,
            product_model, serial_number, purchase_channel, issue_type,
            issue_description, notes, created_date, updated_date,
            completion_date, internal_notes
        FROM after_sales_requests
    """, connection)

    connection.close()

    df.columns = [
        "ID", "Case Number", "Status", "Customer Name", "Email", "Phone",
        "Product Model", "Serial Number", "Purchase Channel / Distributor",
        "Issue Type", "Issue Description", "Customer Notes", "Created Date",
        "Updated Date", "Completion Date", "Internal Notes"
    ]

    export_path = f"after_sales_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    df.to_excel(export_path, index=False)
    format_excel_file(export_path)

    return send_file(export_path, as_attachment=True)


@app.route("/admin")
def admin():
    if "logged_in" not in session:
        return redirect("/login")

    search = request.args.get("search", "")
    deleted = request.args.get("deleted", "")
    updated = request.args.get("updated", "")

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            id, case_number, status, full_name, phone, email,
            country, product_category, product_model, serial_number,
            purchase_date, distributor, created_date, updated_date,
            completion_date, internal_notes
        FROM warranty_registrations
        WHERE
            case_number LIKE ?
            OR full_name LIKE ?
            OR email LIKE ?
            OR serial_number LIKE ?
    """, (
        f"%{search}%",
        f"%{search}%",
        f"%{search}%",
        f"%{search}%"
    ))

    customers = cursor.fetchall()

    cursor.execute("""
        SELECT
            id, case_number, status, customer_name, email, phone,
            product_model, serial_number, purchase_channel, issue_type,
            issue_description, notes, created_date, updated_date,
            completion_date, internal_notes
        FROM after_sales_requests
        WHERE
            case_number LIKE ?
            OR customer_name LIKE ?
            OR email LIKE ?
            OR serial_number LIKE ?
            OR purchase_channel LIKE ?
    """, (
        f"%{search}%",
        f"%{search}%",
        f"%{search}%",
        f"%{search}%",
        f"%{search}%"
    ))

    after_sales_requests = cursor.fetchall()
    connection.close()

    pending_count = 0
    progress_count = 0
    solved_count = 0
    closed_count = 0

    all_cases = list(customers) + list(after_sales_requests)

    for case in all_cases:
        status = case[2]

        if status == "Pending":
            pending_count += 1
        elif status == "In Progress":
            progress_count += 1
        elif status == "Solved":
            solved_count += 1
        elif status == "Closed":
            closed_count += 1

    issue_map = {}

    for request_item in after_sales_requests:
        issue = request_item[9]

        if issue in issue_map:
            issue_map[issue] += 1
        else:
            issue_map[issue] = 1

    issue_labels = list(issue_map.keys())
    issue_counts = list(issue_map.values())

    return render_template(
        "admin.html",
        customers=customers,
        after_sales_requests=after_sales_requests,
        search=search,
        deleted=deleted,
        updated=updated,
        pending_count=pending_count,
        progress_count=progress_count,
        solved_count=solved_count,
        closed_count=closed_count,
        issue_labels=issue_labels,
        issue_counts=issue_counts
    )


@app.route("/add-timeline-update", methods=["POST"])
def add_timeline_update():
    if "logged_in" not in session:
        return redirect("/login")

    case_type = request.form.get("case_type")
    case_number = request.form.get("case_number")
    timeline_update = request.form.get("timeline_update")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    formatted_update = f"{timestamp}\n{timeline_update}\n\n"

    connection = sqlite3.connect("customers.db")
    cursor = connection.cursor()

    if case_type == "warranty":
        cursor.execute("""
            SELECT service_timeline
            FROM warranty_registrations
            WHERE case_number = ?
        """, (case_number,))

        result = cursor.fetchone()
        current_timeline = result[0] if result and result[0] else ""
        updated_timeline = current_timeline + formatted_update

        cursor.execute("""
            UPDATE warranty_registrations
            SET service_timeline = ?,
                updated_date = ?
            WHERE case_number = ?
        """, (updated_timeline, timestamp, case_number))

    elif case_type == "after-sales":
        cursor.execute("""
            SELECT service_timeline
            FROM after_sales_requests
            WHERE case_number = ?
        """, (case_number,))

        result = cursor.fetchone()
        current_timeline = result[0] if result and result[0] else ""
        updated_timeline = current_timeline + formatted_update

        cursor.execute("""
            UPDATE after_sales_requests
            SET service_timeline = ?,
                updated_date = ?
            WHERE case_number = ?
        """, (updated_timeline, timestamp, case_number))

    connection.commit()
    connection.close()

    return redirect(f"/case/{case_type}/{case_number}")


@app.route("/track", methods=["GET", "POST"])
def track_case():
    case_data = None
    not_found = False

    if request.method == "POST":
        case_number = request.form.get("case_number")

        connection = sqlite3.connect("customers.db")
        cursor = connection.cursor()

        if case_number.startswith("W-"):
            cursor.execute("""
                SELECT
                    id, case_number, status, full_name, phone, email,
                    country, product_category, product_model, serial_number,
                    purchase_date, distributor, created_date, updated_date,
                    completion_date, internal_notes
                FROM warranty_registrations
                WHERE case_number = ?
            """, (case_number,))
            case_data = cursor.fetchone()

        elif case_number.startswith("AS-"):
            cursor.execute("""
                SELECT
                    id, case_number, status, customer_name, email, phone,
                    product_model, serial_number, purchase_channel, issue_type,
                    issue_description, notes, created_date, updated_date,
                    completion_date, internal_notes
                FROM after_sales_requests
                WHERE case_number = ?
            """, (case_number,))
            case_data = cursor.fetchone()

        connection.close()

        if not case_data:
            not_found = True

    return render_template(
        "track.html",
        case_data=case_data,
        not_found=not_found
    )
@app.route("/delete-attachment", methods=["POST"])
def delete_attachment():

    if "logged_in" not in session:
        return redirect("/login")

    case_number = request.form.get("case_number")
    filename = request.form.get("filename")

    # handle both old and new attachment paths
    if "/" in filename:
        file_path = os.path.join(
            app.config["UPLOAD_FOLDER"],
            filename
        )
    else:
        file_path = os.path.join(
            app.config["UPLOAD_FOLDER"],
            filename
        )

    if os.path.exists(file_path):
        os.remove(file_path)

    if case_number.startswith("W-"):
        case_type = "warranty"
    else:
        case_type = "after-sales"

    return redirect(f"/case/{case_type}/{case_number}")
@app.route("/manuals/<category>/<filename>")
def download_manual(category, filename):

    folder_path = os.path.join(
        "manuals",
        category
    )

    return send_from_directory(
        folder_path,
        filename,
        as_attachment=True
    )
if __name__ == "__main__":
    app.run(debug=True)
